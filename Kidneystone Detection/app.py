from flask import Flask, render_template, request, redirect, session, url_for
import sqlite3
import os
import pandas as pd
from datetime import datetime
import joblib
import numpy as np
from PIL import Image
import matplotlib.pyplot as plt
import random
from openpyxl import Workbook, load_workbook


app = Flask(__name__)
app.secret_key = "supersecretkey"

UPLOAD_FOLDER = "static/uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ---------------- EXCEL FILES ----------------
USERS_FILE = "users.xlsx"
LOGIN_FILE = "login_history.xlsx"
PREDICTION_FILE = "predictions.xlsx"

def create_file_if_not_exists(filename, headers):
    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.append(headers)
        wb.save(filename)

create_file_if_not_exists(USERS_FILE, ["Username", "Password", "Registered Date"])
create_file_if_not_exists(LOGIN_FILE, ["Username", "Login Time"])
create_file_if_not_exists(PREDICTION_FILE, ["Username", "Date", "Result"])

# ---------------- LOAD MODEL ----------------
model = joblib.load("kidney_rf_model.pkl")

# ---------------- IMAGE PREPROCESS ----------------
def preprocess_image(img_path):
    img = Image.open(img_path).convert("L")
    img = img.resize((50, 50))
    img = np.array(img).flatten() / 255.0
    return img.reshape(1, -1)

# ---------------- DATABASE INIT ----------------
def init_db():
    conn = sqlite3.connect("database.db")
    c = conn.cursor()

    c.execute('''CREATE TABLE IF NOT EXISTS users(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT,
        password TEXT
    )''')

    c.execute('''CREATE TABLE IF NOT EXISTS predictions(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT,
        date TEXT,
        result TEXT
    )''')

    conn.commit()
    conn.close()

init_db()

# ---------------- SOLUTIONS LIST ----------------
solutions_list = [
    "Drink 3-4 liters of water daily",
    "Reduce salt intake",
    "Avoid oxalate-rich foods",
    "Increase citrus fruits",
    "Regular kidney checkups",
    "Limit processed foods"
]

# ---------------- HOME PAGE (BEFORE LOGIN) ----------------
@app.route("/")
def home():
    return render_template("home.html")

# ---------------- REGISTER ----------------
@app.route("/register", methods=["GET","POST"])
def register():

    if request.method == "POST":
        username = request.form["username"]
        password = request.form["password"]
        reg_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        conn = sqlite3.connect("database.db")
        c = conn.cursor()
        c.execute("INSERT INTO users(username,password) VALUES (?,?)",
                  (username,password))
        conn.commit()
        conn.close()

        # Save to Excel
        wb = load_workbook(USERS_FILE)
        ws = wb.active
        ws.append([username, password, reg_date])
        wb.save(USERS_FILE)

        return redirect(url_for("login"))

    return render_template("register.html")

# ---------------- LOGIN ----------------
@app.route("/login", methods=["GET","POST"])
def login():

    if request.method == "POST":

        username = request.form["username"]
        password = request.form["password"]
        login_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        conn = sqlite3.connect("database.db")
        c = conn.cursor()

        c.execute("SELECT * FROM users WHERE username=? AND password=?",
                  (username,password))

        user = c.fetchone()
        conn.close()

        if user:

            session["user"] = username

            # Save login history
            wb = load_workbook(LOGIN_FILE)
            ws = wb.active
            ws.append([username, login_time])
            wb.save(LOGIN_FILE)

            return redirect(url_for("index"))

    return render_template("login.html")

# ---------------- INDEX PAGE (AFTER LOGIN) ----------------
@app.route("/index")
def index():

    if "user" not in session:
        return redirect(url_for("login"))

    return render_template("index.html")

# ---------------- PREDICT ----------------
@app.route("/predict", methods=["GET","POST"])
def predict():

    if "user" not in session:
        return redirect(url_for("login"))

    result = ""
    confidence = ""
    stone_size = ""
    risk = ""
    doctor = ""
    image_path = ""

    if request.method == "POST":

        file = request.files["image"]
        image_path = os.path.join(UPLOAD_FOLDER, file.filename)
        file.save(image_path)

        img = preprocess_image(image_path)

        pred = model.predict(img)[0]
        prob = model.predict_proba(img)[0]

        confidence = round(max(prob) * 100, 2)

        if pred == 1:

            result = "Kidney Stone Detected"

            stone_size = np.random.choice([
                "Small (2-4 mm)",
                "Medium (5-7 mm)",
                "Large (8-12 mm)"
            ])

            risk = np.random.choice([
                "Moderate Risk",
                "High Risk"
            ])

            doctor = np.random.choice([
                "Dr. Ramesh Kumar - Urologist",
                "Dr. Sneha Patil - Nephrologist",
                "Dr. Arjun Mehta - Kidney Specialist"
            ])

        else:

            result = "Normal Kidney"
            stone_size = "No stone detected"
            risk = "Low Risk"
            doctor = "No consultation required"

        date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Save to SQL
        conn = sqlite3.connect("database.db")
        c = conn.cursor()

        c.execute(
            "INSERT INTO predictions(username,date,result) VALUES (?,?,?)",
            (session["user"], date, result)
        )

        conn.commit()
        conn.close()

        # Save to Excel
        wb = load_workbook(PREDICTION_FILE)
        ws = wb.active
        ws.append([session["user"], date, result])
        wb.save(PREDICTION_FILE)

    return render_template(
        "predict.html",
        prediction=result,
        confidence=confidence,
        stone_size=stone_size,
        risk=risk,
        doctor=doctor,
        image=image_path
    )

# ---------------- ANALYSIS ----------------
@app.route("/analysis")
def analysis():

    if "user" not in session:
        return redirect(url_for("login"))

    conn = sqlite3.connect("database.db")
    df = pd.read_sql_query("SELECT result FROM predictions", conn)
    conn.close()

    total = len(df)
    normal_count = 0
    stone_count = 0

    if not df.empty:

        counts = df["result"].value_counts()

        normal_count = counts.get("Normal Kidney", 0)
        stone_count = counts.get("Kidney Stone Detected", 0)

        plt.figure(figsize=(6,4))
        counts.plot(kind="bar")
        plt.title("Prediction Analysis")
        plt.tight_layout()

        plt.savefig("static/bar_chart.png")
        plt.close()

    return render_template(
        "analysis.html",
        total=total,
        normal_count=normal_count,
        stone_count=stone_count
    )

# ---------------- SOLUTIONS ----------------
@app.route("/solutions")
def solutions():

    if "user" not in session:
        return redirect(url_for("login"))

    return render_template(
        "solutions.html",
        solutions=random.sample(solutions_list, 3)
    )

# ---------------- PROJECT PAGE ----------------
@app.route("/project")
def project():
    return render_template("project.html")


# ---------------- ABOUT PAGE ----------------
@app.route("/about")
def about():
    return render_template("about.html")
# ---------------- LOGOUT ----------------
@app.route("/logout")
def logout():

    session.clear()

    return redirect(url_for("home"))

# ---------------- RUN APP ----------------
if __name__ == "__main__":
    app.run(debug=True)